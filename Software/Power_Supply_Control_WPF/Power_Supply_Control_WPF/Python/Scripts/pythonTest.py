from pythonTranslate import *
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import openpyxl.styles
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.table import TableStyleInfo
import numpy as np
import time

log_to_console("""
========================================
 PYTHON DEBUG MODE (PSU SCRIPT)
========================================

This script is waiting for VS Code to attach.

To debug:

1. Open VS Code
2. Open the folder containing this script
3. Go to Run and Debug (Ctrl + Shift + D)
4. Select: "Attach to PSU Python"
5. Press F5

Once attached:
- Execution will continue automatically
- You can set breakpoints in VS Code
- You can inspect variables and step through code

Waiting for debugger connection...
========================================
""")

wait_for_debugger()
log_to_console("Debugger Connected!")

voltage = read_voltage("P")
set_voltage(12.250, "P")
voltage = read_voltage("P")
current = read_current("P")

# Plot Example
log_to_console("Gathering Data...")
new_rows_list = []

# Number of test samples, taken with a delay of SampleDelay (seconds)
n = 1024
SampleDelay = 0.05

start_time = time.perf_counter()

for i in range(n):
    calculated_voltage = read_voltage("P")
    calculated_current = read_current("P")
    time_snapshot = time.perf_counter()
    time_now = time_snapshot - start_time
    new_rows_list.append({
        "Voltage (V)": calculated_voltage,
        "Current (A)": calculated_current,
        "Time (Seconds)": time_now
    })
    time.sleep(SampleDelay)

df = pd.DataFrame(new_rows_list)

# Save spreadsheet
excel_file = "pyReport.xlsx"
df.to_excel(excel_file, index=False, sheet_name="Measurements")

# Generate plot
plt.figure(figsize=(6,4))
plt.plot(df["Time (Seconds)"], df["Voltage (V)"], marker='o', label="Voltage (V)")
plt.xlabel("Time (Seconds)")
plt.ylabel("Volts")
plt.title("(Python Generated Graph)" + '\n' + "Voltage")
plt.grid(True)
plt.legend()
plot_fileV = "pyChartVoltage.png"
plt.savefig(plot_fileV)
plt.close()

plt.figure(figsize=(6,4))
plt.plot(df["Time (Seconds)"], df["Current (A)"], marker='s', label="Current (A)")
plt.xlabel("Time (Seconds)")
plt.ylabel("Amps")
plt.title("(Python Generated Graph)" + '\n' + "Current")
plt.grid(True)
plt.legend()
plot_fileI = "pyChartCurrent.png"
plt.savefig(plot_fileI)
plt.close()

# Insert plot into Excel
wb = load_workbook(excel_file)
ws = wb.active
img = Image(plot_fileV)
ws.add_image(img, "E2")
img = Image(plot_fileI)
ws.add_image(img, "E22")
wb.save(excel_file)
log_to_console("Report generated")

# Statistics Example (On new sheet)
stats = {
    "Duration": np.max(df["Time (Seconds)"]),
    "Mean Voltage": np.mean(df["Voltage (V)"]),
    "Std Dev Voltage": np.std(df["Voltage (V)"]),
    "Min Voltage": np.min(df["Voltage (V)"]),
    "Max Voltage": np.max(df["Voltage (V)"]),
    "Mean Current": np.mean(df["Current (A)"]),
    "Std Dev Current": np.std(df["Current (A)"]),
    "Min Current": np.min(df["Current (A)"]),
    "Max Current": np.max(df["Current (A)"]),
}

stats_df = pd.DataFrame(
    stats.items(),
    columns=["Metric", "Value"]
)

with pd.ExcelWriter(
        excel_file,
        engine="openpyxl",
        mode="a") as writer:

    stats_df.to_excel(
        writer,
        sheet_name="Statistics",
        index=False
    )

    ws = writer.sheets["Statistics"]

    # Header formatting
    header_fill = openpyxl.styles.PatternFill(
        fill_type="solid",
        fgColor="4472C4"
    )

    header_font = openpyxl.styles.Font(
        color="FFFFFF",
        bold=True
    )

    thin_border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style="thin"),
        right=openpyxl.styles.Side(style="thin"),
        top=openpyxl.styles.Side(style="thin"),
        bottom=openpyxl.styles.Side(style="thin")
    )

    # Format all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = openpyxl.styles.Alignment(vertical="center")

    # Format header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Auto-size columns
    for column in ws.columns:
        max_length = 0

        for cell in column:
            if cell.value is not None:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        ws.column_dimensions[column[0].column_letter].width = max_length + 3

    # Format numeric values
    for cell in ws["B"][1:]:
        cell.number_format = "0.000"

    # Create Excel table
    last_row = ws.max_row
    table = Table(
        displayName="StatisticsTable",
        ref=f"A1:B{last_row}"
    )

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showRowStripes=True,
        showColumnStripes=False
    )

    table.tableStyleInfo = style
    ws.add_table(table)

# PSU Test Example
# for voltage in range(2, 15):
#    psu.set_voltage(voltage)
#
#    measured_voltage = psu.read_voltage()
#    measured_current = psu.read_current()
#
#    results.append({
#        "Set Voltage": voltage,
#        "Measured Voltage": measured_voltage,
#        "Measured Current": measured_current
#    })
